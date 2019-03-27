from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.borders import Border, Side
# from openpyxl.utils import coordinate_from_string, column_index_from_string
# from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
# import PIL import Image, ImageOps
import io
import urllib3
import image
import openpyxl
import pandas as pd
import pdb
import sys
import time as t
# import win32com.client
import xlsxwriter
# from docx import Document
# from docx.enum.text import WD_ALIGN_PARAGRAPH

TIMESTAMP = t.strftime("%Y%m%d_%H%M%S")

def main():
    # file_path = sys.argv[1]
    file_path = "C:\\Users\\SBellala\\Desktop\\macro_replacement\\Files\\OnDemandTest\\NIZ_OnDemandExport_erios@skyitgroup_02-05-2019_Bottom_588.xls"
    split_file_path = file_path.split('\\')

    filename = split_file_path[-1]
    client_id = filename.split('_')[0]
    report_type = filename.split('_')[1]
    file_id = filename.split('_')[-1].replace('.xls','')
    tmp_files = []
    # pdb.set_trace()

    if file_path.endswith(".xls") or file_path.endswith(".xlsx"):
       df = pd.read_excel(file_path)

       file = filename[:-4] + "_" + TIMESTAMP +".xlsx"
       writer = pd.ExcelWriter(file,engine='xlsxwriter')
       df.to_excel(writer, sheet_name='StyleSelling', index=False)
       rows = df.shape[0] + 1
       columns = df.shape[1] + 1
       workbook = writer.book
       # workbook.filename = filename[:-4] + "_" + TIMESTAMP + "_" + file_id + ".xlsx"
       workbook.filename = 'test.xlsx'
       temp_file = workbook.filename
       # temp_file = 'test.xlsx' #only for testing
       writer.save()
       writer.close()
       # pdb.set_trace()

       # wb = openpyxl.load_workbook('test.xlsx')
       wb = openpyxl.load_workbook(temp_file)
       ws = wb.active
       #
       # for i in range(2,rows):
       #    pdb.set_trace()
       #    for row in range(i,rows):
       #        if ws.cell(row,1).value == ws.cell(row+1,1).value:
       #            row=row+1#pdb.set_trace()
       #        else:
       #            cell = ws.cell(row=i, column=1)
       #            i = row
       #            pdb.set_trace()
       #            print(cell.value)
       #            break
       pdb.set_trace()
       for i in range(2,rows):
          for j in range(i,rows):
              if ws.cell(j,1).value == ws.cell(j+1,1).value:
                  j=j+1
              else:
                  # ws.merge_cells(start_row = i, start_column = 1, end_row = j, end_column = 1)
                  cell = ws.cell(row=j+1, column=1)
                  print(cell.value)
                  i = j
                  # pdb.set_trace()
                  break

       wb.save(temp_file)


if __name__ == "__main__":
    main()
