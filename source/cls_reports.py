
# import EmailModule as em
import image
import io
import logging
import openpyxl
import os
# import os, os.path,psutil
import pandas as pd
import pdb
import urllib3
import sys
import time as t
# import win32com.client
import xlsxwriter
# from docx import Document
# from docx.enum.text import WD_ALIGN_PARAGRAPH
from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.borders import Border, Side
# from openpyxl.utils import coordinate_from_string, column_index_from_string
# from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
# from PIL import Image, ImageOps


# FILE_DIR = ('\\\\192.168.100.25\\Qlik\\ReportManager_Test\\')
FILE_DIR = "C:\\report_manager\\001files\\"    #for local
# ROOT_DIR = 'D:\\cia_replacement_test\\'
ROOT_DIR = "C:\\report_manager\\"
# ICON_DIR = "D:\\cia_replacement_test\\icons\\"
ICON_DIR = "C:\\report_manager\\icons\\"
TIMESTAMP = t.strftime("%Y%m%d_%H%M%S")
# TMP_FILES = []


def generate_cls_report(file_path, filename, client_id, report_type, file_id, sel_filename):
# def generate_cls_report():

    # pdb.set_trace()
    icon_path = ICON_DIR + client_id + "_Icon.png"
    if file_path.endswith(".xls") or file_path.endswith(".xlsx"):
       df = pd.read_excel(file_path)

       for filename in os.listdir(FILE_DIR + "Selections\\"):
           sel_file_path = FILE_DIR + "Selections\\" + filename
           if filename == sel_filename:
                df_sel = pd.read_excel (sel_file_path, header=None)
                selections = []
                s = []
                selections = df_sel.values.tolist()
                selections = [x for x in selections[0] if str(x) != 'nan']
                s = selections[0].split('\n')
                s = list(filter(None, s))
                s_len = len(s)
                df_sel = pd.DataFrame({'selections': s})

       # file = filename[:-4] + "_" + TIMESTAMP +".xlsx"
       output_file = client_id + '_Output_' + TIMESTAMP + '_' + file_id +'.xlsx'

       writer = pd.ExcelWriter(output_file,engine='xlsxwriter')
       # Infosheet = writer.add_worksheet()
       df_sel.to_excel(writer, sheet_name='Info', startrow=4, startcol=1, header=False, index=False)
       df.to_excel(writer, sheet_name='StyleSelling', index=False)
       rows = df.shape[0] + 1
       columns = df.shape[1] + 1
       workbook = writer.book
       # workbook.filename = filename[:-4] + "_" + TIMESTAMP + "_" + file_id + ".xlsx"
       # workbook.filename = 'test.xlsx'
       workbook.filename = output_file
       temp_file = workbook.filename
       # temp_file = 'test.xlsx' #only for testing
       writer.save()
       # writer.close(test.xlsx)
       # pdb.set_trace()

       # wb = openpyxl.load_workbook('test.xlsx')
       wb = openpyxl.load_workbook(temp_file)
       # active_ws=wb.get_sheet_by_name("Data")
       wb.active = 2
       ws = wb["StyleSelling"]
       # ws = wb.active
       ws.sheet_view.showGridLines = False
       ws.row_dimensions[1].height = 20
       ws.column_dimensions['A'].width = 20.75
       ws.column_dimensions['N'].width = 20.75

       # # pdb.set_trace()
       for col in range(2, columns-1):
           cell_coord = ws.cell(row = 1, column = col).coordinate[0]
           ws.column_dimensions[cell_coord].width = 10.75

       # Updating last cell in image column null
       # pdb.set_trace()

       thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

       thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))

       ver_thick_border = Border(left=Side(style='thick'),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style=None))

       hor_thick_border = Border(left=Side(style=None),
                          right=Side(style=None),
                          top=Side(style='thick'),
                          bottom=Side(style='thick'))

       brc_thick_border = Border(left=Side(style=None),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style='thick'))

       blc_thick_border = Border(left=Side(style=None),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style='thick'))


       # Formatting Column headers
       header_clr = openpyxl.styles.colors.Color(rgb='00336699')
       header_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=header_clr)
       header_font = Font(color='00FFFFFF')
       for col in range(1,columns):
            cell = ws.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText= True)
            cell.fill = header_fill
            cell.font = header_font


       # Merging 'Pattern Name' Column and aligning
       pattern_clr = openpyxl.styles.colors.Color(rgb='00B5E2FF')
       pattern_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=pattern_clr)
       pattern_font = Font(size = 13)
       for i in range(2,rows):
           for j in range(i,rows):
                if ws.cell(row = j, column = 1).value == ws.cell(row = j+1, column = 1).value:
                    j = j+1
                else:
                    ws.merge_cells(start_row = i, start_column = 1, end_row = j, end_column = 1)
                    cell = ws.cell(row=i, column=1)
                    cell.border = thick_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.fill = pattern_fill
                    cell.font = pattern_font
                    i = j
                    break

       # Merging 'Style' Column and aligning
       style_font = Font(size = 8)
       for i in range(2,rows):
           for j in range(i,rows):
                if ws.cell(row = j, column = 2).value == ws.cell(row = j+1, column = 2).value:
                    j = j+1
                else:
                    ws.merge_cells(start_row = i, start_column = 2, end_row = j, end_column = 2)
                    cell = ws.cell(row=i, column=2)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = style_font
                    i = j
                    break

       # formatting color column
       color_font = Font(size = 14)
       for row in range(2,rows):
           cell = ws.cell(row = row, column = 3)
           cell.font = color_font

       # formatting Material_desc column
       material_font = Font(size = 8)
       for row in range(2,rows):
           cell = ws.cell(row = row, column = 4)
           cell.font = material_font

       # formatting numbers
       numbers_font = Font(size = 14)
       for col in range(5,columns-1):
           for row in range(2,rows+3):
               cell = ws.cell(row = row, column = col)
               cell.font = numbers_font


       # Merging 'Image' Column and aligning
       for i in range(2,rows):
           for j in range(i,rows):
                if ws.cell(row = j, column = 14).value == ws.cell(row = j+1, column = 14).value:
                    j = j+1
                else:
                    ws.merge_cells(start_row = i, start_column= 14, end_row = j, end_column = 14)
                    cell = ws.cell(row=i, column=14)
                    cell.border = ver_thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    i = j
                    break

       # Borders for remaining Image columns
       for row in range(2,rows):
           cell = ws.cell(row = row, column = 14)
           cell.border = ver_thick_border


       # inserting Image loop
       for row in range(2, rows):
           cell = ws.cell(row = row, column = 14).value
           cell_coord = ws.cell(row=row, column=14).coordinate
           if cell == None:
               row = row + 1
           else:
               http = urllib3.PoolManager()
               r = http.request('GET', cell)
               image_file = io.BytesIO(r.data)
               img = Image(image_file)
               ws.add_image(img, cell_coord)
               # ws.cell.aligning = WD_ALIGN_PARAGRAPH.CENTER
               ws.cell(row = row, column = 14).value = None

       # updating row height and adding borders to Totals
       # Cells(j, 14).Interior.Color = RGB(255, 255, 255)
       total_clr = openpyxl.styles.colors.Color(rgb='00C4C2C0')
       total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=total_clr)
       total_font = Font(bold=True, size = 14)
       # cell.fill = my_fill
       for row in range(2,rows):
            if ws.cell(row = row, column = 2).value == 'Total':
                ws.row_dimensions[row].height = 20
                for col in range(2, columns-1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = hor_thick_border
                    cell.fill = total_fill
                    cell.font = total_font
            else:
                ws.row_dimensions[row].height = 35

       # updating font of Style total cells
       style_total_font = Font(bold=True, size = 12)
       for row in range(2,rows):
            if ws.cell(row = row, column = 2).value == 'Total':
                cell = ws.cell(row=row, column=2)
                cell.font = style_total_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

       # updating main total row height and adding border
       main_total_font = Font(bold=True, size = 14)
       for row in range(2,rows+1):
            if ws.cell(row = row, column = 1).value == 'Total':
                ws.row_dimensions[row].height = 35
                for col in range(1, columns-1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = hor_thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = main_total_font
            else:
                pass

       # formatting Main total row
       main_tot_font = Font(size = 12, bold = True)
       if ws.cell(row = rows, column = 1).value == 'Total':
           cell = ws.cell(row = rows, column = 1)
           cell.font = main_tot_font
           cell.alignment = Alignment(horizontal='left', vertical='center')
           cell.border = ver_thick_border

       # formatting Main total cell (Image column)
       # brc = botton right corner
       brc_cell = ws.cell(row = rows, column = columns-1)
       brc_cell.border = brc_thick_border
       brc_cell.value = None

       # formatting Main total cell (column 1)
       # blc = botton right corner
       blc_cell = ws.cell(row = rows, column = 1)
       blc_cell.border = thick_border

       # Aligning cells to center
       for column in range(3, columns-1):
           for row in range(2, rows):
               cell = ws.cell(row=row, column=column)
               cell.alignment = Alignment(horizontal='center', vertical='center')

       # format LW ST% column)
       for row in range(2,rows+1):
           for col in range(columns-3,columns):
               cell = ws.cell(row = row, column = col)
               cell.number_format = '0.0%'

       # format LW AUR column (currency))
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 6)
           # cell.number_format = '"$"#,##0.0_-'
           cell.number_format = '"$"#,##0.0_);("$"#,##0.0)'

       # format LW Sales $ (currency))
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 7)
           # cell.number_format = '"$"#,##0_-'
           cell.number_format = '"$"#,##0_);("$"#,##0)'

       # format STD AUR)
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 9)
           # cell.number_format = '"$"#,##0.0_-'
           cell.number_format = '"$"#,##0.0_);("$"#,##0.0)'

       # format STD Sales $
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 10)
           # cell.number_format = '"$"#,##0_-'
           cell.number_format = '"$"#,##0_);("$"#,##0)'

    """ Adding Selections to report """

    wb.active = 1
    ws = wb["Info"]
    # ws = wb.active
    # updating dimensions for rows and columns
    ws.row_dimensions[1].height = 52
    ws.row_dimensions[2].height = 30
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    no_border = Border(left=Side(style=None),
                       right=Side(style=None),
                       top=Side(style=None),
                       bottom=Side(style=None))


    # Merging 1st and 2nd row
    ws.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = 4)
    ws.merge_cells(start_row = 2, start_column = 1, end_row = 2, end_column = 4)

    # formatting style selling report cell
    ss_font = Font(size = 20) #ss --> style selling
    ss_cell = ws.cell(2,1)
    ss_cell.value = 'Style Selling Report'
    ss_cell.font = ss_font
    ss_cell.alignment = Alignment(horizontal='center', vertical='center')

    # formatting Report Generated: cell
    rg_font = Font(bold=True, size = 11) #rg --> Report Generated
    rg_cell = ws.cell(3,1)
    rg_cell.value = 'Report Generated:'
    rg_cell.font = rg_font

    # adding Timestamp   20190314_133248
    date_time = TIMESTAMP[4:6] + '/' + TIMESTAMP[6:8] + '/' + TIMESTAMP[0:4] + ' ' + TIMESTAMP[9:11] + ':' + TIMESTAMP[11:13]
    ws.cell(3,2).value = date_time

    # formatting Selections: cell
    se_font = Font(bold=True, size = 11) #se --> Selections:
    se_cell = ws.cell(5,1)
    se_cell.value = 'Selections:'
    se_cell.font = se_font

    # inserting Client icon
    cell_coord = ws.cell(row=1, column=2).coordinate
    img = openpyxl.drawing.image.Image(icon_path)
    ws.add_image(img, cell_coord)


    wb.save(temp_file)


def email_reports(client_id,filename,report_type):
    #pdb.set_trace()
    from_email = 'support'
    subject = client_id + ' ' + 'StyleSelling' + ' ' + 'Report'
    attachment = client_id + 'StyleSelling.xlsx'
    message = 'Please' + ' ' + 'See' + ' ' + 'Attached.'
    #file_name = temp_file
    #pdb.set_trace()
    if report_type == 'OnDemandExport':
        recipients = filename.split("_")[2] + ".com"
        logging.debug("Sending email")
        em.send_email_from(from_email,[recipients],subject, message,attachment)
    if report_type == 'ScheduledExport':
        recipients = ["sneelakantan@skyitgroup.com"]
        logging.debug("Sending email")
        em.send_email_from(from_email,recipients,subject, message,attachment)


if __name__ == "__main__":

    # pdb.set_trace()
    # file_path = sys.argv[1]
    file_path = "C:\\report_manager\\001files\\CLS_OnDemandExport_sneelakantan@skyitgroup_01-18-2019_275.xls"
    split_file_path = file_path.split('\\')

    filename = split_file_path[-1]
    client_id = filename.split('_')[0]
    report_type = filename.split('_')[1]
    file_id = filename.split('_')[-1].replace('.xls','')
    sel_filename = filename.replace('OnDemandExport','Selections')
    # tmp_files = []

    # logging.basicConfig(filename= ROOT_DIR + 'log\\pythonlog.txt',level=logging.DEBUG)
    logging.basicConfig(filename= ROOT_DIR + 'log\\' + client_id + 'ReportManager_log.txt',level=logging.DEBUG)
    logging.debug('Python file called' + t.strftime("%Y%m%d_%H%M%S"))
    # pdb.set_trace()



    generate_cls_report(file_path, filename, client_id, report_type, file_id, sel_filename)
    # generate_cls_report()
