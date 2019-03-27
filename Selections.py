from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.utils import coordinate_from_string, column_index_from_string
# from openpyxl.utils import get_column_letter
# from openpyxl.drawing.Image import Image
from PIL import Image, ImageOps
import io
import urllib3
import image
import openpyxl
import pandas as pd
import pdb
import sys
import time as t
# import win32com.client
import xlsxwriter
# import System.Linq

TIMESTAMP = t.strftime("%Y%m%d_%H%M%S")

def scale(image, max_size, method=Image.ANTIALIAS):
    """
    resize 'image' to 'max_size' keeping the aspect ratio
    and place it in center of white 'max_size' image
    """
    image.thumbnail(max_size, method)
    offset = (int((max_size[0] - image.size[0]) / 2), int((max_size[1] - image.size[1]) / 2))
    back = Image.new("RGB", max_size, "white")
    back.paste(image, offset)

    return back

def main():
    # file_path = sys.argv[1]
    file_path = "C:\\Users\\SBellala\\Desktop\\macro_replacement\\Files\\OnDemandTest\\CLS_Selections_sneelakantan@skyitgroup_01-18-2019_274.xls"
    icon_path = "C:\\Users\\SBellala\\Desktop\\macro_replacement\\icons\\CLS_Icon.png"
    # icon_path = "C:\\Users\\SBellala\\Desktop\\macro_replacement\\icons\\" + client_id + "_Icon.png"

    split_file_path = file_path.split('\\')

    filename = split_file_path[-1]
    client_id = filename.split('_')[0]
    report_type = filename.split('_')[1]
    file_id = filename.split('_')[-1].replace('.xls','')
    tmp_files = []

    # pdb.set_trace()
    if file_path.endswith(".xls") or file_path.endswith(".xlsx"):
       df = pd.read_excel (file_path, header=None)
       selections = []
       s = []
       selections = df.values.tolist()
       selections = [x for x in selections[0] if str(x) != 'nan']
       s = selections[0].split('\n')
       s = list(filter(None, s))
       s_len = len(s)
       df = pd.DataFrame({'selections': s})

       file = filename[:-4] + "_" + TIMESTAMP +".xlsx"
       writer = pd.ExcelWriter(file,engine='xlsxwriter')
       df.to_excel(writer, sheet_name='Info', startrow=4, startcol=1, header=False, index=False)

       workbook = writer.book
       # workbook.filename = filename[:-4] + "_" + TIMESTAMP + "_" + file_id + ".xlsx"
       workbook.filename = 'test.xlsx'
       temp_file = workbook.filename
       # temp_file = 'test.xlsx' #only for testing
       writer.save()

       wb = openpyxl.load_workbook(temp_file)
       ws = wb.active

       # updating dimensions for rows and columns
       ws.row_dimensions[1].height = 52
       ws.row_dimensions[2].height = 30
       ws.column_dimensions['A'].width = 18
       ws.column_dimensions['B'].width = 18
       ws.column_dimensions['C'].width = 18
       ws.column_dimensions['D'].width = 18

       no_border = Border(left=Side(style=None),
                          right=Side(style=None),
                          top=Side(style=None),
                          bottom=Side(style=None))


       # Merging 1st and 2nd row
       ws.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = 4)
       ws.merge_cells(start_row = 2, start_column = 1, end_row = 2, end_column = 4)

       # formatting style selling report cell
       ss_font = Font(size = 20) #ss --> style selling
       ss_cell = ws.cell(2,1)
       ss_cell.value = 'Style Selling Report'
       ss_cell.font = ss_font
       ss_cell.alignment = Alignment(horizontal='center', vertical='center')

       # formatting Report Generated: cell
       rg_font = Font(bold=True, size = 11) #rg --> Report Generated
       rg_cell = ws.cell(3,1)
       rg_cell.value = 'Report Generated:'
       rg_cell.font = rg_font

       # adding Timestamp   20190314_133248
       date_time = TIMESTAMP[4:6] + '/' + TIMESTAMP[6:8] + '/' + TIMESTAMP[0:4] + ' ' + TIMESTAMP[9:11] + ':' + TIMESTAMP[11:13]
       ws.cell(3,2).value = date_time

       # formatting Selections: cell
       se_font = Font(bold=True, size = 11) #se --> Selections:
       se_cell = ws.cell(5,1)
       se_cell.value = 'Selections:'
       se_cell.font = se_font

       # inserting Client icon
       cell_coord = ws.cell(row=1, column=2).coordinate
       img = openpyxl.drawing.image.Image(icon_path)
       ws.add_image(img, cell_coord)

       # testing padding
       pdb.set_trace()
       cell_coord = ws.cell(row=13, column=2).coordinate
       # img = openpyxl.drawing.image.Image(icon_path)
       img = Image.open(icon_path)
       old_size = img.size   # +size = 180 * 63
       # delta_w = 20
       # delta_h = 10
       # padding = (10, 5, 10, 5 )
       # new_img = ImageOps.expand(img, padding)
       # new_img.show()
       max_size = (200, 73)

       final_image = scale(img, max_size, method=Image.ANTIALIAS)
       # im = openpyxl.drawing.image.Image(final_image)
       # ws.add_image(final_image, cell_coord)
       final_image.show()

       wb.save(temp_file)


if __name__ == "__main__":
    main()
